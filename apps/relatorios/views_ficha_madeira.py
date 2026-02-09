from __future__ import annotations

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import TemplateView

from apps.cadastros.models import Cliente, TipoMadeira
from apps.romaneio.models import ItemRomaneio, Romaneio

from .views_ficha_romaneio import get_mes_ano, _safe_filename  # reutiliza helpers


def _madeiras_queryset(request):
    """
    QuerySet base da Ficha de Madeiras (por item), aplicando filtros e ordenação.

    Filtros (GET):
      - mes, ano (obrigatórios com fallback)
      - cliente (opcional)
      - numero_romaneio (opcional)
      - tipo_romaneio (opcional: NORMAL|COM_FRETE)
      - tipo_madeira_id (opcional: FK -> TipoMadeira)

    Ordenação (GET):
      - sort: data|numero|madeira|tipo|valor_unit|m3|total
      - dir: asc|desc
    """
    mes, ano = get_mes_ano(request)

    cliente_id = (request.GET.get("cliente") or "").strip()
    numero_romaneio = (request.GET.get("numero_romaneio") or "").strip()
    tipo_romaneio = (request.GET.get("tipo_romaneio") or "").strip().upper()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    sort = (request.GET.get("sort") or "data").strip().lower()
    direction = (request.GET.get("dir") or "asc").strip().lower()
    desc = direction == "desc"

    qs = (
        ItemRomaneio.objects.filter(
            romaneio__data_romaneio__month=mes,
            romaneio__data_romaneio__year=ano,
        )
        .select_related("romaneio", "romaneio__cliente", "tipo_madeira")
    )

    # ===== filtros =====
    if cliente_id:
        qs = qs.filter(romaneio__cliente_id=cliente_id)

    if numero_romaneio:
        qs = qs.filter(romaneio__numero_romaneio=numero_romaneio)

    if tipo_romaneio in {"NORMAL", "COM_FRETE"}:
        qs = qs.filter(romaneio__tipo_romaneio=tipo_romaneio)

    if tipo_madeira_id:
        qs = qs.filter(tipo_madeira_id=tipo_madeira_id)

    # ===== ordenação (mapeada para campos do ORM) =====
    sort_map = {
        "data": "romaneio__data_romaneio",
        "numero": "romaneio__numero_romaneio",
        "madeira": "tipo_madeira__nome",
        "tipo": "romaneio__tipo_romaneio",
        "valor_unit": "valor_unitario",
        "m3": "quantidade_m3_total",
        "total": "valor_total",
    }
    field = sort_map.get(sort, "romaneio__data_romaneio")
    if desc:
        field = f"-{field}"

    # tie-breakers para ficar estável
    return qs.order_by(
        field,
        "romaneio__data_romaneio",
        "romaneio__numero_romaneio",
        "tipo_madeira__nome",
        "id",
    )


class RelatorioMadeirasView(LoginRequiredMixin, TemplateView):
    template_name = "relatorios/ficha_madeiras.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        mes, ano = get_mes_ano(self.request)
        qs = _madeiras_queryset(self.request)

        cliente_id = (self.request.GET.get("cliente") or "").strip()

        totais = qs.aggregate(
            total_m3=Sum("quantidade_m3_total"),
            total_itens=Sum("valor_total"),
        )

        context.update({
            "rows": qs,
            "mes": mes,
            "ano": ano,
            "cliente_id": cliente_id or "",
            "clientes": Cliente.objects.filter(ativo=True).order_by("nome"),
            "tipos_madeira": TipoMadeira.objects.order_by("nome"),
            "meses": range(1, 13),
            "anos": [d.year for d in Romaneio.objects.dates("data_romaneio", "year", order="ASC")] or [timezone.localdate().year],
            "total_m3": totais["total_m3"] or 0,
            "total_itens": totais["total_itens"] or 0,
            "sort": (self.request.GET.get("sort") or "data"),
            "dir": (self.request.GET.get("dir") or "asc"),
        })
        return context


@login_required
def ficha_madeiras_export_excel(request):
    """Exporta a Ficha de Madeiras (ItemRomaneio) para Excel, respeitando filtros/ordenação."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    mes, ano = get_mes_ano(request)
    qs = _madeiras_queryset(request)

    cliente_id = (request.GET.get("cliente") or "").strip()
    cliente_nome = "Todos"
    if cliente_id:
        c = Cliente.objects.filter(pk=cliente_id).first()
        if c:
            cliente_nome = c.nome

    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()
    tipo_madeira_nome = "Todas"
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            tipo_madeira_nome = tm.nome

    brand_fill = PatternFill("solid", fgColor="246B29")
    head_fill = PatternFill("solid", fgColor="EEF3EF")
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")
    total_fill = PatternFill("solid", fgColor="FFF3CD")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold = Font(bold=True)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, widths: dict[int, float]):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb = Workbook()
    ws = wb.active
    ws.title = f"Madeiras {mes:02d}-{ano}"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"FICHA DE MADEIRAS — {mes:02d}/{ano}"
    ws["A1"].font = title_font
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Cliente: {cliente_nome}  |  Madeira: {tipo_madeira_nome}"
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A2"].fill = brand_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4
    headers = ["Data Romaneio", "Nº Romaneio", "Tipo Madeira", "Tipo", "Valor Unitário", "M³", "Total"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = ws["A5"]
    ws.auto_filter.ref = f"A{header_row}:G{header_row}"

    row = header_row + 1
    for idx, item in enumerate(qs, start=1):
        r = item.romaneio

        ws.cell(row=row, column=1, value=r.data_romaneio).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=2, value=r.numero_romaneio)
        ws.cell(row=row, column=3, value=item.tipo_madeira.nome if item.tipo_madeira else "")
        ws.cell(row=row, column=4, value=("Com frete" if r.tipo_romaneio == "COM_FRETE" else "Normal"))
        ws.cell(row=row, column=5, value=float(item.valor_unitario or 0)).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=6, value=float(item.quantidade_m3_total or 0)).number_format = "0.000"
        ws.cell(row=row, column=7, value=float(item.valor_total or 0)).number_format = '"R$" #,##0.00'

        for cidx in range(1, 8):
            cell = ws.cell(row=row, column=cidx)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if cidx in (1, 3, 4) else "right", vertical="center")
            if idx % 2 == 0:
                cell.fill = zebra_fill

        row += 1

    last_row = row - 1
    total_row = last_row + 2

    ws.cell(row=total_row, column=4, value="TOTAL").font = bold
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).border = border
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")

    if last_row >= header_row + 1:
        ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{last_row})").number_format = "0.000"
        ws.cell(row=total_row, column=7, value=f"=SUM(G{header_row+1}:G{last_row})").number_format = '"R$" #,##0.00'
    else:
        ws.cell(row=total_row, column=6, value=0).number_format = "0.000"
        ws.cell(row=total_row, column=7, value=0).number_format = '"R$" #,##0.00'

    for cidx in (6, 7):
        ws.cell(row=total_row, column=cidx).font = bold
        ws.cell(row=total_row, column=cidx).fill = total_fill
        ws.cell(row=total_row, column=cidx).border = border
        ws.cell(row=total_row, column=cidx).alignment = Alignment(horizontal="right")

    set_col_width(ws, {1: 14, 2: 12, 3: 30, 4: 12, 5: 16, 6: 10, 7: 14})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _safe_filename(f"ficha_madeiras_{mes:02d}_{ano}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def ficha_madeiras_export_pdf(request):
    """Exporta a Ficha de Madeiras (ItemRomaneio) para PDF via WeasyPrint, respeitando filtros/ordenação."""
    mes, ano = get_mes_ano(request)
    qs = _madeiras_queryset(request)

    cliente_id = (request.GET.get("cliente") or "").strip()
    cliente_nome = "Todos"
    if cliente_id:
        c = Cliente.objects.filter(pk=cliente_id).first()
        if c:
            cliente_nome = c.nome

    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()
    tipo_madeira_nome = "Todas"
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            tipo_madeira_nome = tm.nome

    context = {
        "rows": list(qs),
        "mes": mes,
        "ano": ano,
        "cliente_nome": cliente_nome,
        "tipo_madeira_nome": tipo_madeira_nome,
        "total_m3": qs.aggregate(s=Sum("quantidade_m3_total"))["s"] or 0,
        "total_itens": qs.aggregate(s=Sum("valor_total"))["s"] or 0,
        "now": timezone.localtime(),
    }

    html_string = render_to_string("relatorios/ficha_madeiras_pdf.html", context)
    base_url = request.build_absolute_uri("/")

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        return HttpResponse(
            f"Falha ao gerar PDF. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    filename = _safe_filename(f"ficha_madeiras_{mes:02d}_{ano}.pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response