from __future__ import annotations

import csv
import re

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import IntegerField, Sum
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import ListView

from apps.cadastros.models import Cliente, TipoMadeira
from apps.romaneio.models import Romaneio


def get_mes_ano(request) -> tuple[int, int]:
    """Lê mes/ano da querystring e retorna defaults coerentes."""
    now = timezone.localdate()
    try:
        mes = int(request.GET.get("mes", now.month))
    except (TypeError, ValueError):
        mes = now.month
    try:
        ano = int(request.GET.get("ano", now.year))
    except (TypeError, ValueError):
        ano = now.year
    return mes, ano


def _safe_filename(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    return value or "arquivo"


def _romaneios_queryset(request):
    """
    Queryset base para Ficha de Romaneios (por ROMANEIO):
    - filtros: mes/ano, cliente, numero_romaneio, tipo_madeira_id
    - ordenação: sort/dir (inclui ordenação numérica correta do Nº Romaneio no Postgres)

    Observação:
      - Filtro por tipo de madeira usa relação: Romaneio -> itens -> tipo_madeira.
      - Ao filtrar por itens, é necessário distinct() para não duplicar romaneios.
    """
    mes, ano = get_mes_ano(request)
    cliente_id = (request.GET.get("cliente") or "").strip()
    numero_romaneio = (request.GET.get("numero_romaneio") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    sort = (request.GET.get("sort") or "data").strip().lower()
    direction = (request.GET.get("dir") or "asc").strip().lower()
    desc = direction == "desc"

    qs = Romaneio.objects.select_related("cliente", "motorista").filter(
        data_romaneio__month=mes,
        data_romaneio__year=ano,
    )

    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)

    if numero_romaneio:
        qs = qs.filter(numero_romaneio=numero_romaneio)

    if tipo_madeira_id:
        # Filtra romaneios que possuem ao menos 1 item com a madeira selecionada
        qs = qs.filter(itens__tipo_madeira_id=tipo_madeira_id)

    # Evita duplicar romaneios quando o join em itens retornar múltiplas linhas
    if tipo_madeira_id:
        qs = qs.distinct()

    # Ordenação numérica do Nº Romaneio (quando o campo é CharField)
    # No Postgres, Cast falha se existir valor não-numérico.
    qs = qs.annotate(numero_int=Cast("numero_romaneio", output_field=IntegerField()))

    sort_map = {
        "cliente": "cliente__nome",
        "numero": "numero_int",
        "data": "data_romaneio",
        "m3": "m3_total",
        "total": "valor_total",
    }
    field = sort_map.get(sort, "data_romaneio")
    if desc:
        field = f"-{field}"

    # tie-breakers estáveis (não “brigam” com o primeiro critério)
    return qs.order_by(field, "data_romaneio", "numero_int", "id")


class RelatorioRomaneiosView(LoginRequiredMixin, ListView):
    model = Romaneio
    template_name = "relatorios/ficha_romaneios.html"
    context_object_name = "romaneios"
    paginate_by = 50

    def get_queryset(self):
        return _romaneios_queryset(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)

        context["mes"] = mes
        context["ano"] = ano
        context["clientes"] = Cliente.objects.filter(ativo=True).order_by("nome")
        context["tipos_madeira"] = TipoMadeira.objects.order_by("nome")
        context["meses"] = range(1, 13)

        anos = [d.year for d in Romaneio.objects.dates("data_romaneio", "year", order="ASC")]
        context["anos"] = anos or [timezone.localdate().year]

        # Totais do período (mesmos filtros da listagem)
        rom_qs = _romaneios_queryset(self.request).order_by()  # remove order_by (não afeta filtros)

        totais = rom_qs.aggregate(
            total_m3=Sum("m3_total"),
            total_valor_liquido=Sum("valor_total"),
            total_valor_bruto=Sum("valor_bruto"),
        )
        context["total_m3_periodo"] = totais["total_m3"] or 0
        context["total_valor_periodo"] = totais["total_valor_liquido"] or 0
        context["total_valor_bruto_periodo"] = totais["total_valor_bruto"] or 0

        context["sort"] = (self.request.GET.get("sort") or "data")
        context["dir"] = (self.request.GET.get("dir") or "asc")

        return context


@login_required
def ficha_romaneios_export(request):
    """
    Exporta relatório de romaneios POR ITEM em CSV.
    Respeita filtros: mes/ano, cliente, numero_romaneio, tipo_madeira_id.
    """
    mes, ano = get_mes_ano(request)
    cliente_id = (request.GET.get("cliente") or "").strip()
    numero_romaneio = (request.GET.get("numero_romaneio") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    qs = Romaneio.objects.filter(data_romaneio__month=mes, data_romaneio__year=ano)

    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    if numero_romaneio:
        qs = qs.filter(numero_romaneio=numero_romaneio)
    if tipo_madeira_id:
        qs = qs.filter(itens__tipo_madeira_id=tipo_madeira_id).distinct()

    qs = qs.select_related("cliente", "motorista").prefetch_related("itens__tipo_madeira")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="relatorio_romaneios_{mes:02d}_{ano}.csv"'
    writer = csv.writer(response, delimiter=";")

    writer.writerow([
        "Modalidade",
        "Tipo Romaneio",
        "Nº Romaneio",
        "Data",
        "Cliente",
        "Motorista",
        "Espécie",
        "Qtd Item (m³)",
        "Valor Unit. (R$/m³)",
        "Total Item (R$)",
    ])

    total_m3_geral = 0.0
    total_valor_itens = 0.0

    for romaneio in qs:
        itens = romaneio.itens.all()
        if tipo_madeira_id:
            itens = [i for i in itens if str(getattr(i, "tipo_madeira_id", "")) == str(tipo_madeira_id)]

        for item in itens:
            qtd = float(item.quantidade_m3_total or 0)
            total_item = float(item.valor_total or 0)

            writer.writerow([
                romaneio.get_modalidade_display() if hasattr(romaneio, "get_modalidade_display") else (romaneio.modalidade or ""),
                romaneio.get_tipo_romaneio_display() if hasattr(romaneio, "get_tipo_romaneio_display") else (romaneio.tipo_romaneio or ""),
                romaneio.numero_romaneio,
                romaneio.data_romaneio.strftime("%d/%m/%Y"),
                romaneio.cliente.nome if romaneio.cliente else "",
                romaneio.motorista.nome if romaneio.motorista else "",
                item.tipo_madeira.nome if item.tipo_madeira else "",
                f"{qtd:.3f}",
                f"{float(item.valor_unitario or 0):.2f}",
                f"{total_item:.2f}",
            ])

            total_m3_geral += qtd
            total_valor_itens += total_item

    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "TOTAL", f"{total_m3_geral:.3f}", "", f"{total_valor_itens:.2f}"])
    return response


@login_required
def ficha_romaneios_export_excel(request):
    """Exporta a Ficha de Romaneios (por ROMANEIO) para Excel, respeitando filtros e ordenação."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    mes, ano = get_mes_ano(request)
    cliente_id = (request.GET.get("cliente") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    qs = _romaneios_queryset(request)

    brand_fill = PatternFill("solid", fgColor="246B29")
    head_fill = PatternFill("solid", fgColor="EEF3EF")
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")
    total_fill = PatternFill("solid", fgColor="FFF3CD")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold = Font(bold=True)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, widths: dict[int, float]):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb = Workbook()
    ws = wb.active
    ws.title = f"Romaneios {mes:02d}-{ano}"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"FICHA DE ROMANEIOS — {mes:02d}/{ano}"
    ws["A1"].font = title_font
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    cliente_nome = "Todos"
    if cliente_id:
        c = Cliente.objects.filter(pk=cliente_id).first()
        if c:
            cliente_nome = c.nome

    madeira_nome = "Todas"
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            madeira_nome = tm.nome

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Cliente: {cliente_nome}  |  Madeira: {madeira_nome}"
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A2"].fill = brand_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4
    headers = ["Data", "Nº Romaneio", "Cliente", "Motorista", "Tipo", "M³", "Total (R$)"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.freeze_panes = ws["A5"]
    ws.auto_filter.ref = f"A{header_row}:G{header_row}"

    row = header_row + 1
    for idx, r in enumerate(qs, start=1):
        ws.cell(row=row, column=1, value=r.data_romaneio)
        ws.cell(row=row, column=2, value=r.numero_romaneio)
        ws.cell(row=row, column=3, value=r.cliente.nome if r.cliente else "")
        ws.cell(row=row, column=4, value=r.motorista.nome if r.motorista else "")
        ws.cell(row=row, column=5, value=("Com frete" if r.tipo_romaneio == "COM_FRETE" else "Normal"))
        ws.cell(row=row, column=6, value=float(r.m3_total or 0))
        ws.cell(row=row, column=7, value=float(r.valor_total or 0))

        ws.cell(row=row, column=1).number_format = "dd/mm/yyyy"
        ws.cell(row=row, column=6).number_format = "0.000"
        ws.cell(row=row, column=7).number_format = '"R$" #,##0.00'

        for cidx in range(1, 8):
            cell = ws.cell(row=row, column=cidx)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if cidx in (1, 3, 4, 5) else "right", vertical="center")
            if idx % 2 == 0:
                cell.fill = zebra_fill

        row += 1

    last_row = row - 1
    total_row = last_row + 2

    ws.cell(row=total_row, column=5, value="TOTAL").font = bold
    ws.cell(row=total_row, column=5).fill = total_fill
    ws.cell(row=total_row, column=5).border = border
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row, column=6, value=f"=SUM(F{header_row+1}:F{last_row})")
    ws.cell(row=total_row, column=6).number_format = "0.000"
    ws.cell(row=total_row, column=6).font = bold
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).border = border
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row, column=7, value=f"=SUM(G{header_row+1}:G{last_row})")
    ws.cell(row=total_row, column=7).number_format = '"R$" #,##0.00'
    ws.cell(row=total_row, column=7).font = bold
    ws.cell(row=total_row, column=7).fill = total_fill
    ws.cell(row=total_row, column=7).border = border
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")

    set_col_width(ws, {1: 12, 2: 14, 3: 34, 4: 22, 5: 12, 6: 10, 7: 14})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _safe_filename(f"ficha_romaneios_{mes:02d}_{ano}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def ficha_romaneios_export_pdf(request):
    """Exporta a Ficha de Romaneios (por ROMANEIO) para PDF via WeasyPrint, respeitando filtros e ordenação."""
    mes, ano = get_mes_ano(request)
    cliente_id = (request.GET.get("cliente") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    qs = _romaneios_queryset(request)

    cliente_nome = "Todos"
    if cliente_id:
        c = Cliente.objects.filter(pk=cliente_id).first()
        if c:
            cliente_nome = c.nome

    madeira_nome = "Todas"
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            madeira_nome = tm.nome

    totais = qs.order_by().aggregate(total_m3=Sum("m3_total"), total_valor=Sum("valor_total"))

    context = {
        "rows": list(qs),
        "mes": mes,
        "ano": ano,
        "cliente_nome": cliente_nome,
        "madeira_nome": madeira_nome,
        "total_m3": totais["total_m3"] or 0,
        "total_valor": totais["total_valor"] or 0,
        "now": timezone.localtime(),
    }

    html_string = render_to_string("relatorios/ficha_romaneios_pdf.html", context)
    base_url = request.build_absolute_uri("/")

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        return HttpResponse(
            f"Falha ao gerar PDF. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    filename = _safe_filename(f"ficha_romaneios_{mes:02d}_{ano}.pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def romaneio_export_excel(request, romaneio_id: int):
    """Exporta UM romaneio para Excel (.xlsx) com layout profissional."""
    # (mantive seu código original daqui pra baixo)
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    romaneio = (
        Romaneio.objects.select_related("cliente", "motorista", "usuario_cadastro")
        .prefetch_related("itens__tipo_madeira", "itens__unidades")
        .get(pk=romaneio_id)
    )

    brand_fill = PatternFill("solid", fgColor="246B29")
    head_fill = PatternFill("solid", fgColor="EEF3EF")
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")
    total_fill = PatternFill("solid", fgColor="FFF3CD")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold = Font(bold=True)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, widths: dict[int, float]):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def apply_border_range(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = border

    wb = Workbook()
    ws = wb.active
    ws.title = f"Romaneio {romaneio.numero_romaneio}"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"ROMANEIO Nº {romaneio.numero_romaneio}"
    ws["A1"].font = title_font
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Cliente: {romaneio.cliente.nome if romaneio.cliente else ''}   |   "
        f"Data: {romaneio.data_romaneio.strftime('%d/%m/%Y') if romaneio.data_romaneio else ''}   |   "
        f"Tipo: {romaneio.get_tipo_romaneio_display()}   |   "
        f"Modalidade: {romaneio.get_modalidade_display()}"
    )
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A2"].fill = brand_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws["A4"] = "Resumo"
    ws["A4"].font = Font(bold=True, size=12)
    ws.merge_cells("A4:F4")

    resumo = [
        ("Cliente", romaneio.cliente.nome if romaneio.cliente else ""),
        ("Motorista", romaneio.motorista.nome if romaneio.motorista else "—"),
        ("Cadastrado por", (romaneio.usuario_cadastro.get_full_name() if romaneio.usuario_cadastro else "—")),
        ("Cadastro em", romaneio.data_cadastro.strftime("%d/%m/%Y %H:%M") if getattr(romaneio, "data_cadastro", None) else "—"),
        ("Atualização", romaneio.data_atualizacao.strftime("%d/%m/%Y %H:%M") if getattr(romaneio, "data_atualizacao", None) else "—"),
        ("Total m³", float(romaneio.m3_total or 0)),
        ("Valor Bruto (R$)", float(romaneio.valor_bruto or 0)),
        ("Valor Líquido (R$)", float(romaneio.valor_total or 0)),
    ]

    r0 = 5
    for i, (k, v) in enumerate(resumo):
        row = r0 + i
        ws["A" + str(row)] = k
        ws["A" + str(row)].font = bold
        ws["A" + str(row)].alignment = Alignment(horizontal="right")
        ws["A" + str(row)].fill = PatternFill("solid", fgColor="F2F2F2")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        cell_v = ws.cell(row=row, column=2, value=v)
        cell_v.alignment = Alignment(horizontal="left")

        if k == "Total m³":
            cell_v.number_format = "0.000"
        if "R$" in k:
            cell_v.number_format = '"R$" #,##0.00'

    apply_border_range(ws, min_row=5, max_row=5 + len(resumo) - 1, min_col=1, max_col=6)

    start_row = 5 + len(resumo) + 2
    ws["A" + str(start_row)] = "Itens do Romaneio"
    ws["A" + str(start_row)].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)

    header_row = start_row + 1
    headers = ["Espécie", "Qtd (m³)", "Valor Unit. (R$/m³)", "Total Item (R$)"]
    for col_idx, text in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=text)
        c.font = Font(bold=True, color="1F2937")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    row = header_row + 1
    for idx, item in enumerate(romaneio.itens.all(), start=1):
        especie = item.tipo_madeira.nome if item.tipo_madeira else ""
        qtd = float(item.quantidade_m3_total or 0)
        vu = float(item.valor_unitario or 0)
        vt = float(item.valor_total or 0)

        ws.cell(row=row, column=1, value=especie)
        ws.cell(row=row, column=2, value=qtd)
        ws.cell(row=row, column=3, value=vu)
        ws.cell(row=row, column=4, value=vt)

        for cidx in range(1, 5):
            cell = ws.cell(row=row, column=cidx)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="left" if cidx == 1 else "right")
            if idx % 2 == 0:
                cell.fill = zebra_fill

        ws.cell(row=row, column=2).number_format = "0.000"
        ws.cell(row=row, column=3).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=4).number_format = '"R$" #,##0.00'
        row += 1

    last_item_row = row - 1

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = border

    ws.cell(row=total_row, column=2, value=f"=SUM(B{header_row+1}:B{last_item_row})")
    ws.cell(row=total_row, column=2).number_format = "0.000"
    ws.cell(row=total_row, column=2).font = bold
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).border = border

    ws.cell(row=total_row, column=4, value=f"=SUM(D{header_row+1}:D{last_item_row})")
    ws.cell(row=total_row, column=4).number_format = '"R$" #,##0.00'
    ws.cell(row=total_row, column=4).font = bold
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).border = border

    ws.freeze_panes = ws["A" + str(header_row + 1)]
    ws.auto_filter.ref = f"A{header_row}:D{last_item_row}"

    set_col_width(ws, {1: 38, 2: 12, 3: 20, 4: 16, 5: 2, 6: 2})

    # Aba Unidades (se DETALHADO)
    if romaneio.modalidade == "DETALHADO":
        ws2 = wb.create_sheet("Unidades (Toras)")
        ws2.freeze_panes = "A2"

        ws2.append(["Item/Espécie", "Nº", "Comprimento (m)", "Rôdo (cm)", "Desc. 1 (cm)", "Desc. 2 (cm)", "Qtd (m³)"])
        for col in range(1, 8):
            c = ws2.cell(row=1, column=col)
            c.font = Font(bold=True)
            c.fill = head_fill
            c.border = border
            c.alignment = Alignment(horizontal="center")

        out_row = 2
        for item in romaneio.itens.all():
            especie = item.tipo_madeira.nome if item.tipo_madeira else ""
            for idx, u in enumerate(item.unidades.all(), start=1):
                ws2.cell(row=out_row, column=1, value=especie)
                ws2.cell(row=out_row, column=2, value=idx)
                ws2.cell(row=out_row, column=3, value=float(u.comprimento or 0))
                ws2.cell(row=out_row, column=4, value=float(u.rodo or 0))
                ws2.cell(row=out_row, column=5, value=float(u.desconto_1 or 0))
                ws2.cell(row=out_row, column=6, value=float(u.desconto_2 or 0))
                ws2.cell(row=out_row, column=7, value=float(u.quantidade_m3 or 0))

                for col in range(1, 8):
                    cell = ws2.cell(row=out_row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left" if col == 1 else "right")
                    if out_row % 2 == 0:
                        cell.fill = zebra_fill

                ws2.cell(row=out_row, column=3).number_format = "0.00"
                ws2.cell(row=out_row, column=4).number_format = "0.00"
                ws2.cell(row=out_row, column=5).number_format = "0.00"
                ws2.cell(row=out_row, column=6).number_format = "0.00"
                ws2.cell(row=out_row, column=7).number_format = "0.000"
                out_row += 1

        set_col_width(ws2, {1: 28, 2: 6, 3: 16, 4: 12, 5: 12, 6: 12, 7: 12})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _safe_filename(f"romaneio_{romaneio.numero_romaneio}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def romaneio_export_pdf(request, romaneio_id: int):
    """Exporta UM romaneio para PDF (impressão) usando WeasyPrint."""
    romaneio = get_object_or_404(
        Romaneio.objects.select_related("cliente", "motorista", "usuario_cadastro").prefetch_related(
            "itens__tipo_madeira", "itens__unidades"
        ),
        pk=romaneio_id,
    )

    # Garante consistência dos dados antes de imprimir (SIMPLES e DETALHADO)
    # - SIMPLES: item.atualizar_totais usa quantidade_m3_total informada
    # - DETALHADO: item.atualizar_totais soma unidades
    for item in romaneio.itens.all():
        item.atualizar_totais(save=True, atualizar_romaneio=False)

    romaneio.atualizar_totais(save=True)

    context = {
        "romaneio": romaneio,
        "itens": list(romaneio.itens.all()),
        "now": timezone.localtime(),
    }

    html_string = render_to_string("relatorios/romaneio_pdf.html", context)
    base_url = request.build_absolute_uri("/")

    try:
        from weasyprint import HTML
    except Exception as exc:
        return HttpResponse(
            f"WeasyPrint não está disponível neste servidor. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    try:
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        return HttpResponse(
            f"Falha ao gerar PDF. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    filename = _safe_filename(f"romaneio_{romaneio.numero_romaneio}.pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response