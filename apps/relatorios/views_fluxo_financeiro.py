from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from typing import Iterable

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import TemplateView

from apps.cadastros.models import TipoMadeira, Cliente
from apps.financeiro.models import Pagamento
from apps.romaneio.models import Romaneio

from .views_ficha_romaneio import _safe_filename, get_mes_ano


# =============================================================================
# Tipos / helpers
# =============================================================================
@dataclass(frozen=True)
class MovimentoFluxo:
    """
    Representa uma movimentação no fluxo financeiro.

    Regras de exibição (template):
      - VENDA (romaneio): numero_romaneio/m3/total preenchidos; credito vazio
      - PAGAMENTO: credito preenchido; numero_romaneio/m3/total vazios

    saldo_atual: saldo do cliente APÓS a movimentação (pode ser negativo).
    """
    data: timezone.datetime | timezone.date
    cliente_id: int
    cliente_nome: str

    # venda (romaneio)
    numero_romaneio: str | None = None
    m3: Decimal | None = None
    total: Decimal | None = None

    # pagamento
    credito: Decimal | None = None

    saldo_atual: Decimal = Decimal("0.00")


def _to_decimal(value) -> Decimal:
    try:
        return (value if isinstance(value, Decimal) else Decimal(str(value or 0)))
    except Exception:
        return Decimal("0")


def _calc_saldos_por_movimento(movs: list[MovimentoFluxo]) -> list[MovimentoFluxo]:
    """
    Calcula saldo_atual por cliente, em ordem cronológica (asc).
    Convenção:
      - VENDA aumenta a dívida do cliente => saldo fica mais negativo (saldo -= total)
      - PAGAMENTO reduz a dívida => saldo aumenta (saldo += credito)

    Retorna lista com saldo_atual preenchido, na mesma ordem de entrada.
    """
    saldos: dict[int, Decimal] = {}

    out: list[MovimentoFluxo] = []
    for m in movs:
        saldo_atual = saldos.get(m.cliente_id, Decimal("0.00"))

        if m.total is not None:
            saldo_atual = (saldo_atual - _to_decimal(m.total)).quantize(Decimal("0.01"))
        elif m.credito is not None:
            saldo_atual = (saldo_atual + _to_decimal(m.credito)).quantize(Decimal("0.01"))

        saldos[m.cliente_id] = saldo_atual
        out.append(
            MovimentoFluxo(
                data=m.data,
                cliente_id=m.cliente_id,
                cliente_nome=m.cliente_nome,
                numero_romaneio=m.numero_romaneio,
                m3=m.m3,
                total=m.total,
                credito=m.credito,
                saldo_atual=saldo_atual,
            )
        )
    return out


# =============================================================================
# Querysets e montagem das movimentações
# =============================================================================
def _fluxo_querysets(request):
    mes, ano = get_mes_ano(request)

    numero_romaneio = (request.GET.get("numero_romaneio") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()
    cliente_id = (request.GET.get("cliente_id") or "").strip()  # <-- NOVO

    vendas_qs = (
        Romaneio.objects.filter(
            data_romaneio__month=mes,
            data_romaneio__year=ano,
        )
        .select_related("cliente", "motorista")
    )

    if cliente_id:
        vendas_qs = vendas_qs.filter(cliente_id=cliente_id)

    if numero_romaneio:
        vendas_qs = vendas_qs.filter(numero_romaneio=numero_romaneio)

    if tipo_madeira_id:
        vendas_qs = vendas_qs.filter(itens__tipo_madeira_id=tipo_madeira_id).distinct()

    pagamentos_qs = (
        Pagamento.objects.filter(
            data_pagamento__month=mes,
            data_pagamento__year=ano,
        )
        .select_related("cliente")
    )

    if cliente_id:
        pagamentos_qs = pagamentos_qs.filter(cliente_id=cliente_id)

    if numero_romaneio or tipo_madeira_id:
        cliente_ids = list(vendas_qs.values_list("cliente_id", flat=True).distinct())
        pagamentos_qs = pagamentos_qs.filter(cliente_id__in=cliente_ids)

    return vendas_qs, pagamentos_qs


def _build_movimentacoes(vendas_qs: Iterable[Romaneio], pagamentos_qs: Iterable[Pagamento]) -> list[MovimentoFluxo]:
    """
    Constrói lista única de movimentações (vendas + pagamentos) com saldo por linha.

    Ordenação:
      - Para calcular saldo corretamente, processamos em ordem ASC por data.
      - Para exibir, retornamos em ordem DESC por data (mais recente primeiro).
    """
    movs: list[MovimentoFluxo] = []

    for r in vendas_qs:
        movs.append(
            MovimentoFluxo(
                data=r.data_romaneio,
                cliente_id=r.cliente_id,
                cliente_nome=(r.cliente.nome if r.cliente else ""),
                numero_romaneio=str(r.numero_romaneio),
                m3=_to_decimal(getattr(r, "m3_total", None)),
                total=_to_decimal(getattr(r, "valor_total", None)),
                credito=None,
            )
        )

    for p in pagamentos_qs:
        movs.append(
            MovimentoFluxo(
                data=p.data_pagamento,
                cliente_id=p.cliente_id,
                cliente_nome=(p.cliente.nome if p.cliente else ""),
                numero_romaneio=None,
                m3=None,
                total=None,
                credito=_to_decimal(getattr(p, "valor", None)),
            )
        )

    # ASC para calcular saldo por linha
    movs.sort(key=lambda x: (x.data, x.cliente_nome, (x.numero_romaneio or "")))

    movs_com_saldo = _calc_saldos_por_movimento(movs)

    # DESC para exibição
    movs_com_saldo.sort(key=lambda x: (x.data, x.cliente_nome, (x.numero_romaneio or "")))
    return movs_com_saldo


# =============================================================================
# View (HTML)
# =============================================================================
class RelatorioFluxoView(LoginRequiredMixin, TemplateView):
    template_name = "relatorios/fluxo_financeiro.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)

        vendas_qs, pagamentos_qs = _fluxo_querysets(self.request)

        vendas_total = vendas_qs.aggregate(total=Sum("valor_total")).get("total") or 0
        pagamentos_total = pagamentos_qs.aggregate(total=Sum("valor")).get("total") or 0

        saldo_mes = pagamentos_total - vendas_total
        if saldo_mes > 0:
            saldo_mes_classe = "text-success"
        elif saldo_mes < 0:
            saldo_mes_classe = "text-danger"
        else:
            saldo_mes_classe = "text-secondary"

        # anos disponíveis
        anos_romaneios = [d.year for d in Romaneio.objects.dates("data_romaneio", "year", order="ASC")]
        anos_pagamentos = [d.year for d in Pagamento.objects.dates("data_pagamento", "year", order="ASC")]
        anos = sorted(set(anos_romaneios + anos_pagamentos)) or [timezone.localdate().year]

        movimentacoes = _build_movimentacoes(vendas_qs, pagamentos_qs)

        context.update(
            {
                "mes": mes,
                "ano": ano,
                "meses": range(1, 13),
                "anos": anos,
                "clientes": Cliente.objects.filter(ativo=True).order_by("nome"),
                "cliente_id": (self.request.GET.get("cliente_id") or "").strip(),
                "tipos_madeira": TipoMadeira.objects.order_by("nome"),
                # KPIs
                "vendas": vendas_total,
                "pagamentos": pagamentos_total,
                "saldo_mes": saldo_mes,
                "saldo_mes_classe": saldo_mes_classe,
                # Nova tabela única
                "movimentacoes": movimentacoes,
                # Mantém dados para filtros (template atual usa isso)
                "numero_romaneio": (self.request.GET.get("numero_romaneio") or "").strip(),
                "tipo_madeira_id": (self.request.GET.get("tipo_madeira_id") or "").strip(),
                "now": timezone.localtime(),
            }
        )
        return context


# =============================================================================
# Exports
# =============================================================================
@login_required
def fluxo_financeiro_export_excel(request):
    """
    Exporta o Fluxo Financeiro do período em Excel.
    Agora exporta uma aba única "Movimentações" (compras + pagamentos) + "Resumo".
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    mes, ano = get_mes_ano(request)
    vendas_qs, pagamentos_qs = _fluxo_querysets(request)

    vendas_total = vendas_qs.aggregate(total=Sum("valor_total")).get("total") or 0
    pagamentos_total = pagamentos_qs.aggregate(total=Sum("valor")).get("total") or 0
    saldo = pagamentos_total - vendas_total

    movimentacoes = _build_movimentacoes(vendas_qs, pagamentos_qs)

    brand_fill = PatternFill("solid", fgColor="246B29")
    head_fill = PatternFill("solid", fgColor="EEF3EF")
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold = Font(bold=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, widths: dict[int, float]):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"FLUXO FINANCEIRO — {mes:02d}/{ano}"
    ws["A1"].font = title_font
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    numero_romaneio = (request.GET.get("numero_romaneio") or "").strip()
    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()

    madeira_nome = "Todas"
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            madeira_nome = tm.nome

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Filtro Nº Romaneio: {numero_romaneio if numero_romaneio else '��'}  |  Madeira: {madeira_nome}"
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A2"].fill = brand_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    resumo = [
        ("Vendas (R$)", float(vendas_total or 0)),
        ("Pagamentos (R$)", float(pagamentos_total or 0)),
        ("Saldo (R$)", float(saldo or 0)),
    ]
    start = 4
    for i, (k, v) in enumerate(resumo):
        r = start + i
        ws.cell(row=r, column=1, value=k).font = bold
        ws.cell(row=r, column=2, value=v).number_format = '"R$" #,##0.00'
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border

    set_col_width(ws, {1: 22, 2: 18, 3: 2, 4: 2})

    # Aba Movimentações (única)
    ws_m = wb.create_sheet("Movimentações")
    ws_m.append(["Data", "Nº Romaneio", "Cliente", "M³", "Total (R$)", "Crédito (R$)", "Saldo atual (R$)"])

    for c in range(1, 8):
        cell = ws_m.cell(row=1, column=c)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for idx, m in enumerate(movimentacoes, start=1):
        ws_m.cell(row=row, column=1, value=m.data).number_format = "dd/mm/yyyy"
        ws_m.cell(row=row, column=2, value=m.numero_romaneio or "")
        ws_m.cell(row=row, column=3, value=m.cliente_nome)

        ws_m.cell(row=row, column=4, value=float(m.m3) if m.m3 is not None else None).number_format = "0.000"
        ws_m.cell(row=row, column=5, value=float(m.total) if m.total is not None else None).number_format = '"R$" #,##0.00'
        ws_m.cell(row=row, column=6, value=float(m.credito) if m.credito is not None else None).number_format = '"R$" #,##0.00'
        ws_m.cell(row=row, column=7, value=float(m.saldo_atual)).number_format = '"R$" #,##0.00'

        for c in range(1, 8):
            cell = ws_m.cell(row=row, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if c == 3 else "right", vertical="center")
            if idx % 2 == 0:
                cell.fill = zebra_fill

        row += 1

    ws_m.freeze_panes = "A2"
    ws_m.auto_filter.ref = f"A1:G{max(1, row - 1)}"
    set_col_width(ws_m, {1: 12, 2: 14, 3: 38, 4: 10, 5: 16, 6: 16, 7: 18})

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _safe_filename(f"fluxo_financeiro_{mes:02d}_{ano}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def fluxo_financeiro_export_pdf(request):
    """
    Exporta o Fluxo Financeiro do período para PDF via WeasyPrint.
    Agora exporta uma lista única de movimentações (compras + pagamentos).
    """
    mes, ano = get_mes_ano(request)
    vendas_qs, pagamentos_qs = _fluxo_querysets(request)

    vendas_total = vendas_qs.aggregate(total=Sum("valor_total")).get("total") or 0
    pagamentos_total = pagamentos_qs.aggregate(total=Sum("valor")).get("total") or 0
    saldo = pagamentos_total - vendas_total

    tipo_madeira_id = (request.GET.get("tipo_madeira_id") or "").strip()
    madeira_nome = ""
    if tipo_madeira_id:
        tm = TipoMadeira.objects.filter(pk=tipo_madeira_id).first()
        if tm:
            madeira_nome = tm.nome

    movimentacoes = _build_movimentacoes(vendas_qs, pagamentos_qs)

    context = {
        "mes": mes,
        "ano": ano,
        "vendas_total": vendas_total,
        "pagamentos_total": pagamentos_total,
        "saldo": saldo,
        "movimentacoes": movimentacoes,
        "numero_romaneio": (request.GET.get("numero_romaneio") or "").strip(),
        "madeira_nome": madeira_nome,
        "now": timezone.localtime(),
    }

    html_string = render_to_string("relatorios/fluxo_financeiro_pdf.html", context)
    base_url = request.build_absolute_uri("/")

    try:
        from weasyprint import HTML

        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        return HttpResponse(
            f"Falha ao gerar PDF. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    filename = _safe_filename(f"fluxo_financeiro_{mes:02d}_{ano}.pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response