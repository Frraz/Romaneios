from __future__ import annotations

import csv
import re

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic import ListView, TemplateView

from apps.cadastros.models import Cliente
from apps.financeiro.models import Pagamento
from apps.romaneio.models import ItemRomaneio, Romaneio

def get_mes_ano(request):
    """
    Lê mes/ano da querystring e retorna defaults coerentes.
    """
    now = timezone.localdate()
    try:
        mes = int(request.GET.get("mes", now.month))
    except (TypeError, ValueError):
        mes = now.month
    try:
        ano = int(request.GET.get("ano", now.year))
    except (TypeError, ValueError):
        ano = now.year
    return mes, ano


class RelatorioRomaneiosView(LoginRequiredMixin, ListView):
    model = Romaneio
    template_name = "relatorios/ficha_romaneios.html"
    context_object_name = "romaneios"
    paginate_by = 50

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("cliente", "motorista")
        )
        mes, ano = get_mes_ano(self.request)
        cliente_id = self.request.GET.get("cliente")

        qs = qs.filter(data_romaneio__month=mes, data_romaneio__year=ano)
        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)

        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)
        cliente_id = self.request.GET.get("cliente")

        context["mes"] = mes
        context["ano"] = ano
        context["clientes"] = Cliente.objects.filter(ativo=True).order_by("nome")

        # Anos disponíveis sem varrer tudo em Python
        anos = [d.year for d in Romaneio.objects.dates("data_romaneio", "year", order="ASC")]
        context["anos"] = anos or [timezone.localdate().year]
        context["meses"] = range(1, 13)

        # Totais do período (usando o mesmo filtro do queryset)
        rom_qs = Romaneio.objects.filter(data_romaneio__month=mes, data_romaneio__year=ano)
        if cliente_id:
            rom_qs = rom_qs.filter(cliente_id=cliente_id)

        totais = rom_qs.aggregate(
            total_m3=Sum("m3_total"),
            total_valor_liquido=Sum("valor_total"),
            total_valor_bruto=Sum("valor_bruto"),
        )
        context["total_m3_periodo"] = totais["total_m3"] or 0
        context["total_valor_periodo"] = totais["total_valor_liquido"] or 0
        context["total_valor_bruto_periodo"] = totais["total_valor_bruto"] or 0

        return context


def ficha_romaneios_export(request):
    """
    Exporta relatório de romaneios POR ITEM em CSV.

    Corrigido para refletir o seu model atual:
    - ItemRomaneio não tem comprimento/rodo diretamente (isso fica em UnidadeRomaneio).
    - A quantidade do item é `quantidade_m3_total`.
    """
    mes, ano = get_mes_ano(request)
    cliente_id = request.GET.get("cliente")

    qs = Romaneio.objects.filter(data_romaneio__month=mes, data_romaneio__year=ano)
    if cliente_id:
        qs = qs.filter(cliente_id=cliente_id)
    qs = qs.select_related("cliente", "motorista").prefetch_related("itens__tipo_madeira")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="relatorio_romaneios_{mes:02d}_{ano}.csv"'

    # BOM: Excel pt-BR costuma gostar de ; como delimiter. Se quiser manter vírgula, troque.
    writer = csv.writer(response, delimiter=";")

    writer.writerow([
        "Modalidade",
        "Tipo Romaneio",
        "Nº Romaneio",
        "Data",
        "Cliente",
        "Motorista",
        "Espécie",
        "Qtd Item (m³)",
        "Valor Unit. (R$/m³)",
        "Total Item (R$)",
    ])

    total_m3_geral = 0.0
    total_valor_itens = 0.0

    for romaneio in qs:
        itens = romaneio.itens.all()
        for item in itens:
            qtd = float(item.quantidade_m3_total or 0)
            total_item = float(item.valor_total or 0)

            writer.writerow([
                romaneio.get_modalidade_display() if hasattr(romaneio, "get_modalidade_display") else (romaneio.modalidade or ""),
                romaneio.get_tipo_romaneio_display() if hasattr(romaneio, "get_tipo_romaneio_display") else (romaneio.tipo_romaneio or ""),
                romaneio.numero_romaneio,
                romaneio.data_romaneio.strftime("%d/%m/%Y"),
                romaneio.cliente.nome if romaneio.cliente else "",
                romaneio.motorista.nome if romaneio.motorista else "",
                item.tipo_madeira.nome if item.tipo_madeira else "",
                f"{qtd:.3f}",
                f"{float(item.valor_unitario or 0):.2f}",
                f"{total_item:.2f}",
            ])

            total_m3_geral += qtd
            total_valor_itens += total_item

    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "TOTAL", f"{total_m3_geral:.3f}", "", f"{total_valor_itens:.2f}"])
    return response


@login_required
def romaneio_export_excel(request, romaneio_id: int):
    """
    Exporta UM romaneio para Excel (.xlsx) com layout profissional:
    - Cabeçalho estilizado
    - Resumo
    - Itens com formatação numérica/moeda, filtro, freeze panes e total
    - (Se DETALHADO) uma aba adicional com Unidades (Toras)
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    romaneio = (
        Romaneio.objects
        .select_related("cliente", "motorista", "usuario_cadastro")
        .prefetch_related("itens__tipo_madeira", "itens__unidades")
        .get(pk=romaneio_id)
    )

    # ===== Helpers de estilo =====
    brand_fill = PatternFill("solid", fgColor="246B29")   # verde
    head_fill = PatternFill("solid", fgColor="EEF3EF")    # cabeçalho tabela (claro)
    zebra_fill = PatternFill("solid", fgColor="F7F7F7")   # linhas alternadas
    total_fill = PatternFill("solid", fgColor="FFF3CD")   # total (amarelo claro)

    white_bold = Font(color="FFFFFF", bold=True, size=14)
    title_font = Font(bold=True, size=16, color="FFFFFF")
    bold = Font(bold=True)
    small_muted = Font(color="666666", size=10)

    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_col_width(ws, widths: dict[int, float]):
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def apply_border_range(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = border

    # ===== Workbook =====
    wb = Workbook()
    ws = wb.active
    ws.title = f"Romaneio {romaneio.numero_romaneio}"

    # ===== Cabeçalho =====
    ws.merge_cells("A1:F1")
    ws["A1"] = f"ROMANEIO Nº {romaneio.numero_romaneio}"
    ws["A1"].font = title_font
    ws["A1"].fill = brand_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Cliente: {romaneio.cliente.nome if romaneio.cliente else ''}   |   "
        f"Data: {romaneio.data_romaneio.strftime('%d/%m/%Y') if romaneio.data_romaneio else ''}   |   "
        f"Tipo: {romaneio.get_tipo_romaneio_display()}   |   "
        f"Modalidade: {romaneio.get_modalidade_display()}"
    )
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=11)
    ws["A2"].fill = brand_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ===== Resumo =====
    # Bloco 2 colunas (Label / Valor)
    ws["A4"] = "Resumo"
    ws["A4"].font = Font(bold=True, size=12)
    ws.merge_cells("A4:F4")

    resumo = [
        ("Cliente", romaneio.cliente.nome if romaneio.cliente else ""),
        ("Motorista", romaneio.motorista.nome if romaneio.motorista else "—"),
        ("Cadastrado por", (romaneio.usuario_cadastro.get_full_name() if romaneio.usuario_cadastro else "—")),
        ("Cadastro em", romaneio.data_cadastro.strftime("%d/%m/%Y %H:%M") if getattr(romaneio, "data_cadastro", None) else "—"),
        ("Atualização", romaneio.data_atualizacao.strftime("%d/%m/%Y %H:%M") if getattr(romaneio, "data_atualizacao", None) else "—"),
        ("Total m³", float(romaneio.m3_total or 0)),
        ("Valor Bruto (R$)", float(romaneio.valor_bruto or 0)),
        ("Valor Líquido (R$)", float(romaneio.valor_total or 0)),
    ]

    r0 = 5
    for i, (k, v) in enumerate(resumo):
        row = r0 + i
        ws["A" + str(row)] = k
        ws["A" + str(row)].font = bold
        ws["A" + str(row)].alignment = Alignment(horizontal="right")
        ws["A" + str(row)].fill = PatternFill("solid", fgColor="F2F2F2")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        cell_v = ws.cell(row=row, column=2, value=v)
        cell_v.alignment = Alignment(horizontal="left")

        # formatos
        if k == "Total m³":
            cell_v.number_format = "0.000"
        if "R$" in k:
            cell_v.number_format = '"R$" #,##0.00'

    apply_border_range(ws, min_row=5, max_row=5 + len(resumo) - 1, min_col=1, max_col=6)

    # ===== Tabela de itens =====
    start_row = 5 + len(resumo) + 2  # espaço após resumo

    ws["A" + str(start_row)] = "Itens do Romaneio"
    ws["A" + str(start_row)].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)

    header_row = start_row + 1
    headers = ["Espécie", "Qtd (m³)", "Valor Unit. (R$/m³)", "Total Item (R$)"]
    for col_idx, text in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=text)
        c.font = Font(bold=True, color="1F2937")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    ws.row_dimensions[header_row].height = 18

    row = header_row + 1
    for idx, item in enumerate(romaneio.itens.all(), start=1):
        especie = item.tipo_madeira.nome if item.tipo_madeira else ""
        qtd = float(item.quantidade_m3_total or 0)
        vu = float(item.valor_unitario or 0)
        vt = float(item.valor_total or 0)

        ws.cell(row=row, column=1, value=especie)
        ws.cell(row=row, column=2, value=qtd)
        ws.cell(row=row, column=3, value=vu)
        ws.cell(row=row, column=4, value=vt)

        # estilos
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", horizontal="left" if c == 1 else "right")
            if idx % 2 == 0:
                cell.fill = zebra_fill

        ws.cell(row=row, column=2).number_format = "0.000"
        ws.cell(row=row, column=3).number_format = '"R$" #,##0.00'
        ws.cell(row=row, column=4).number_format = '"R$" #,##0.00'
        row += 1

    last_item_row = row - 1

    # Linha total (com fórmula)
    total_row = row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).border = border

    ws.cell(row=total_row, column=2, value=f"=SUM(B{header_row+1}:B{last_item_row})")
    ws.cell(row=total_row, column=2).number_format = "0.000"
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).border = border

    ws.cell(row=total_row, column=4, value=f"=SUM(D{header_row+1}:D{last_item_row})")
    ws.cell(row=total_row, column=4).number_format = '"R$" #,##0.00'
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).border = border

    # Preencher células intermediárias do total com estilo
    for c in (3,):
        cell = ws.cell(row=total_row, column=c, value="")
        cell.fill = total_fill
        cell.border = border

    # Ajustes úteis do Excel
    ws.freeze_panes = ws["A" + str(header_row + 1)]  # congela até o cabeçalho
    ws.auto_filter.ref = f"A{header_row}:D{last_item_row}"

    # Larguras
    set_col_width(ws, {
        1: 38,  # espécie
        2: 12,  # qtd
        3: 20,  # valor unit
        4: 16,  # total
        5: 2,
        6: 2,
    })

    # Impressão (opcional, mas deixa profissional)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"

    # ===== Aba opcional: Unidades (Toras) =====
    if romaneio.modalidade == "DETALHADO":
        ws2 = wb.create_sheet("Unidades (Toras)")
        ws2.freeze_panes = "A2"

        ws2.append(["Romaneio", romaneio.numero_romaneio, "Cliente", romaneio.cliente.nome if romaneio.cliente else ""])
        ws2.append(["Item/Espécie", "Nº", "Comprimento (m)", "Rôdo (cm)", "Desc. 1 (cm)", "Desc. 2 (cm)", "Qtd (m³)"])

        # Header style
        for col in range(1, 8):
            c = ws2.cell(row=2, column=col)
            c.font = Font(bold=True)
            c.fill = head_fill
            c.border = border
            c.alignment = Alignment(horizontal="center")

        out_row = 3
        for item in romaneio.itens.all():
            especie = item.tipo_madeira.nome if item.tipo_madeira else ""
            for idx, u in enumerate(item.unidades.all(), start=1):
                ws2.cell(row=out_row, column=1, value=especie)
                ws2.cell(row=out_row, column=2, value=idx)
                ws2.cell(row=out_row, column=3, value=float(u.comprimento or 0))
                ws2.cell(row=out_row, column=4, value=float(u.rodo or 0))
                ws2.cell(row=out_row, column=5, value=float(u.desconto_1 or 0))
                ws2.cell(row=out_row, column=6, value=float(u.desconto_2 or 0))
                ws2.cell(row=out_row, column=7, value=float(u.quantidade_m3 or 0))

                for col in range(1, 8):
                    cell = ws2.cell(row=out_row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left" if col == 1 else "right")
                    if out_row % 2 == 0:
                        cell.fill = zebra_fill

                ws2.cell(row=out_row, column=3).number_format = "0.00"
                ws2.cell(row=out_row, column=4).number_format = "0.00"
                ws2.cell(row=out_row, column=5).number_format = "0.00"
                ws2.cell(row=out_row, column=6).number_format = "0.00"
                ws2.cell(row=out_row, column=7).number_format = "0.000"
                out_row += 1

        set_col_width(ws2, {1: 28, 2: 6, 3: 16, 4: 12, 5: 12, 6: 12, 7: 12})

    # ===== Response =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = _safe_filename(f"romaneio_{romaneio.numero_romaneio}.xlsx")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _safe_filename(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    return value or "arquivo"


@login_required
def romaneio_export_pdf(request, romaneio_id: int):
    """
    Exporta UM romaneio para PDF (impressão) usando WeasyPrint.
    Recomendado para produção em Ubuntu VPS.
    """
    romaneio = get_object_or_404(
        Romaneio.objects.select_related(
            "cliente",
            "motorista",
            "usuario_cadastro",
        ).prefetch_related(
            "itens__tipo_madeira",
            "itens__unidades",
        ),
        pk=romaneio_id,
    )

    context = {
        "romaneio": romaneio,
        "itens": list(romaneio.itens.all()),
        "now": timezone.localtime(),
    }

    html_string = render_to_string("relatorios/romaneio_pdf.html", context)
    base_url = request.build_absolute_uri("/")

    try:
        from weasyprint import HTML  # import local (mais robusto no deploy)
    except Exception as exc:
        # Retorna erro legível (e evita 500 "cego")
        return HttpResponse(
            f"WeasyPrint não está disponível neste servidor. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    try:
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
    except Exception as exc:
        # Ajuda MUITO no deploy: você vê o erro no response/log
        return HttpResponse(
            f"Falha ao gerar PDF. Erro: {exc}",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    filename = _safe_filename(f"romaneio_{romaneio.numero_romaneio}.pdf")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response

class RelatorioMadeirasView(LoginRequiredMixin, TemplateView):
    template_name = "relatorios/ficha_madeiras.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)

        relatorio = (
            ItemRomaneio.objects.filter(
                romaneio__data_romaneio__month=mes,
                romaneio__data_romaneio__year=ano,
            )
            .values("tipo_madeira__nome")
            .annotate(
                total_m3=Sum("quantidade_m3_total"),
                total_valor=Sum("valor_total"),
            )
            .order_by("-total_m3")
        )

        context["relatorio"] = relatorio
        context["mes"] = mes
        context["ano"] = ano
        context["meses"] = range(1, 13)
        context["anos"] = [d.year for d in Romaneio.objects.dates("data_romaneio", "year", order="ASC")] or [timezone.localdate().year]
        return context


class RelatorioFluxoView(LoginRequiredMixin, TemplateView):
    template_name = "relatorios/fluxo_financeiro.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)

        vendas = (
            Romaneio.objects.filter(data_romaneio__month=mes, data_romaneio__year=ano)
            .aggregate(total=Sum("valor_total"))
            .get("total")
            or 0
        )

        pagamentos = (
            Pagamento.objects.filter(data_pagamento__month=mes, data_pagamento__year=ano)
            .aggregate(total=Sum("valor"))
            .get("total")
            or 0
        )

        saldo_mes = pagamentos - vendas
        saldo_mes_classe = "text-secondary"
        if saldo_mes > 0:
            saldo_mes_classe = "text-success"
        elif saldo_mes < 0:
            saldo_mes_classe = "text-danger"

        context.update({
            "saldo_mes_classe": saldo_mes_classe,
            "mes": mes,
            "ano": ano,
            "vendas": vendas,
            "pagamentos": pagamentos,
            "saldo_mes": saldo_mes,
            "vendas_detalhadas": Romaneio.objects.filter(
                data_romaneio__month=mes, data_romaneio__year=ano
            ).select_related("cliente", "motorista"),
            "pagamentos_detalhados": Pagamento.objects.filter(
                data_pagamento__month=mes, data_pagamento__year=ano
            ).select_related("cliente"),
        })
        return context


class RelatorioSaldoClientesView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = "relatorios/saldo_clientes.html"
    context_object_name = "clientes"

    def get_queryset(self):
        tipo_saldo = self.request.GET.get("tipo_saldo", "todos")

        qs = Cliente.objects.all()

        # Se saldo_atual for property calculada em Python, isso não dá para filtrar no banco.
        # Mantive o comportamento original (lista em Python), mas filtrando ativos pode ser útil.
        clientes = list(qs)

        if tipo_saldo == "negativos":
            clientes = [c for c in clientes if c.saldo_atual < 0]
        elif tipo_saldo == "positivos":
            clientes = [c for c in clientes if c.saldo_atual > 0]
        elif tipo_saldo == "zerados":
            clientes = [c for c in clientes if c.saldo_atual == 0]

        return sorted(clientes, key=lambda c: c.saldo_atual)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tipos_saldo"] = ["todos", "negativos", "positivos", "zerados"]
        return context


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "relatorios/dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        mes, ano = get_mes_ano(self.request)

        romaneios_mes = Romaneio.objects.filter(data_romaneio__month=mes, data_romaneio__year=ano)
        totais_mes = romaneios_mes.aggregate(
            total_m3=Sum("m3_total"),
            total_valor=Sum("valor_total"),
            qtd_romaneios=Count("id"),
        )

        context["mes"] = mes
        context["ano"] = ano
        context["total_m3_mes"] = totais_mes["total_m3"] or 0
        context["total_faturado_mes"] = totais_mes["total_valor"] or 0
        context["qtd_romaneios_mes"] = totais_mes["qtd_romaneios"] or 0

        todos_clientes = Cliente.objects.all()
        saldos_negativos = [c.saldo_atual for c in todos_clientes if c.saldo_atual < 0]
        context["saldo_total_receber"] = abs(sum(saldos_negativos)) if saldos_negativos else 0

        devedores = sorted(
            [c for c in todos_clientes if c.saldo_atual < 0],
            key=lambda c: c.saldo_atual,
        )[:5]
        context["maiores_devedores"] = devedores

        top_clientes = (
            romaneios_mes.values("cliente__nome")
            .annotate(total_comprado=Sum("valor_total"))
            .order_by("-total_comprado")[:5]
        )
        context["top_clientes_mes"] = top_clientes

        vendas_por_madeira = (
            ItemRomaneio.objects.filter(
                romaneio__data_romaneio__month=mes,
                romaneio__data_romaneio__year=ano,
            )
            .values("tipo_madeira__nome")
            .annotate(
                total_m3=Sum("quantidade_m3_total"),
                total_valor=Sum("valor_total"),
            )
            .order_by("-total_m3")[:10]
        )
        context["vendas_por_madeira"] = vendas_por_madeira

        return context